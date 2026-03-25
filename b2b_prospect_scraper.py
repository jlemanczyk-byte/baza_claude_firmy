#!/usr/bin/env python3
"""
B2B Prospect Database Builder
Scrapes Polish job portals and business directories to find B2B companies
that recently posted sales-related job offers.

Sources: pracuj.pl, indeed.pl, panoramafirm.pl, rejestr.io
Output: Excel file with prospect data and statistics.

Usage:
    python b2b_prospect_scraper.py [--output FILE] [--min-prospects N]
"""

import argparse
import logging
import re
import sys
import time
from collections import Counter
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from typing import Optional
from urllib.parse import quote_plus, urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SEARCH_POSITIONS = [
    "handlowiec",
    "specjalista ds. sprzedaży",
    "przedstawiciel handlowy",
    "key account manager",
    "new business manager",
    "BDM",
    "sales manager",
    "opiekun klienta",
    "kierownik sprzedaży",
]

TARGET_INDUSTRIES = [
    "produkcja przemysłowa",
    "produkcja",
    "dystrybucja",
    "IT",
    "software",
    "FMCG",
    "logistyka",
    "usługi B2B",
    "e-commerce",
    "handel",
    "technologia",
    "przemysł",
]

EXCLUDED_INDUSTRIES = [
    "OZE",
    "odnawialne źródła energii",
    "fotowoltaika",
    "finanse",
    "ubezpieczenia",
    "bankowość",
    "bank",
    "ubezpieczenie",
    "finansowy",
    "kredyt",
    "pożyczka",
    "inwestycje",
    "giełda",
    "fundusze",
    "broker",
    "windykacja",
]

DECISION_MAKER_TITLES_PRIORITY = {
    1: [
        "prezes", "ceo", "właściciel", "founder", "dyrektor generalny",
        "managing director", "prezes zarządu", "współwłaściciel",
    ],
    2: [
        "dyrektor sprzedaży", "sales director", "vp sales",
        "wiceprezes ds. sprzedaży", "head of sales", "cso",
        "chief sales officer", "dyrektor handlowy",
        "dyrektor ds. sprzedaży", "kierownik działu sprzedaży",
    ],
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "pl-PL,pl;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

REQUEST_TIMEOUT = 15
DELAY_BETWEEN_REQUESTS = 2.0  # seconds – be polite

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class Prospect:
    company_name: str = ""
    industry: str = ""
    website: str = ""
    city: str = ""
    decision_maker_name: str = ""
    decision_maker_title: str = ""
    decision_maker_priority: int = 0
    email: str = ""
    phone: str = ""
    job_title: str = ""
    job_portal: str = ""
    job_date: str = ""
    job_url: str = ""
    is_active: bool = True
    nip: str = ""
    krs: str = ""


# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------

_session = requests.Session()
_session.headers.update(HEADERS)


def fetch(url: str, retries: int = 2) -> Optional[BeautifulSoup]:
    """Fetch URL and return parsed BeautifulSoup, or None on failure."""
    for attempt in range(retries + 1):
        try:
            time.sleep(DELAY_BETWEEN_REQUESTS)
            resp = _session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            resp.raise_for_status()
            return BeautifulSoup(resp.text, "lxml")
        except requests.RequestException as exc:
            log.warning("Fetch %s attempt %d failed: %s", url, attempt + 1, exc)
            if attempt < retries:
                time.sleep(2 ** attempt)
    return None


def fetch_text(url: str) -> Optional[str]:
    """Fetch raw text content from URL."""
    try:
        time.sleep(DELAY_BETWEEN_REQUESTS)
        resp = _session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        resp.raise_for_status()
        return resp.text
    except requests.RequestException:
        return None


# ---------------------------------------------------------------------------
# Industry filtering
# ---------------------------------------------------------------------------

def is_excluded(text: str) -> bool:
    """Return True if text matches excluded industries."""
    text_lower = text.lower()
    return any(excl.lower() in text_lower for excl in EXCLUDED_INDUSTRIES)


def classify_industry(text: str) -> str:
    """Try to match text against target industries. Return best match or ''."""
    text_lower = text.lower()
    if is_excluded(text_lower):
        return ""
    for ind in TARGET_INDUSTRIES:
        if ind.lower() in text_lower:
            return ind
    return ""


# ---------------------------------------------------------------------------
# Contact extraction helpers
# ---------------------------------------------------------------------------

EMAIL_RE = re.compile(
    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}",
)
PHONE_RE = re.compile(
    r"(?:\+48[\s\-]?)?(?:\d[\s\-]?){9}",
)
NIP_RE = re.compile(r"\bNIP[\s:]*(\d[\d\s\-]{8,12}\d)\b", re.IGNORECASE)
KRS_RE = re.compile(r"\bKRS[\s:]*(\d{10})\b", re.IGNORECASE)


def extract_emails(text: str) -> list[str]:
    found = EMAIL_RE.findall(text)
    blocked = {"example.com", "example.pl", "domena.pl", "email.com"}
    return [e for e in found if e.split("@")[1].lower() not in blocked]


def extract_phones(text: str) -> list[str]:
    raw = PHONE_RE.findall(text)
    cleaned = []
    for p in raw:
        digits = re.sub(r"\D", "", p)
        if len(digits) == 9:
            digits = "+48" + digits
        elif len(digits) == 11 and digits.startswith("48"):
            digits = "+" + digits
        if len(digits) >= 11:
            cleaned.append(digits)
    return cleaned


# ---------------------------------------------------------------------------
# Scraper: Pracuj.pl
# ---------------------------------------------------------------------------

def scrape_pracuj(position: str, max_pages: int = 3) -> list[Prospect]:
    """Scrape pracuj.pl for job offers matching *position*."""
    prospects = []
    base = "https://www.pracuj.pl/praca/{query};kw"
    query = quote_plus(position)
    url = base.format(query=query)

    log.info("Pracuj.pl: searching '%s'", position)

    for page in range(1, max_pages + 1):
        page_url = f"{url}?pn={page}" if page > 1 else url
        soup = fetch(page_url)
        if not soup:
            break

        # Pracuj.pl renders job offers in various div structures
        offers = soup.select(
            "div[data-test='section-offers'] div[data-test='default-offer'], "
            "div.listing_b1hnba8, "
            "div[class*='offer-details'], "
            "a[data-test='link-offer']"
        )
        if not offers:
            # Fallback: look for any anchor with job-related href
            offers = soup.find_all("a", href=re.compile(r"/praca/.*,oferta,"))

        for offer in offers:
            p = Prospect(job_portal="pracuj.pl")

            # Title
            title_el = (
                offer.select_one("h2[data-test='offer-title'] a")
                or offer.select_one("h2 a")
                or offer.select_one("a")
            )
            if title_el:
                p.job_title = title_el.get_text(strip=True)
                href = title_el.get("href", "")
                if href and not href.startswith("http"):
                    href = urljoin("https://www.pracuj.pl", href)
                p.job_url = href
            else:
                text = offer.get_text(strip=True)[:120]
                p.job_title = text

            # Company
            company_el = (
                offer.select_one("[data-test='text-company-name']")
                or offer.select_one("h3")
                or offer.select_one("span.listing_sj2yyt0")
            )
            if company_el:
                p.company_name = company_el.get_text(strip=True)

            # Location
            loc_el = (
                offer.select_one("[data-test='text-region']")
                or offer.select_one("span[class*='location']")
            )
            if loc_el:
                p.city = loc_el.get_text(strip=True).split(",")[0].strip()

            if p.company_name and not is_excluded(p.company_name):
                prospects.append(p)

        log.info("  page %d → %d offers so far", page, len(prospects))

    return prospects


# ---------------------------------------------------------------------------
# Scraper: Indeed.pl
# ---------------------------------------------------------------------------

def scrape_indeed(position: str, max_pages: int = 3) -> list[Prospect]:
    """Scrape indeed.com/jobs (Poland) for job offers."""
    prospects = []
    query = quote_plus(position)
    base = f"https://pl.indeed.com/jobs?q={query}&l=Polska"

    log.info("Indeed.pl: searching '%s'", position)

    for page in range(max_pages):
        page_url = f"{base}&start={page * 10}" if page > 0 else base
        soup = fetch(page_url)
        if not soup:
            break

        cards = soup.select(
            "div.job_seen_beacon, "
            "div.jobsearch-SerpJobCard, "
            "div[class*='result'], "
            "td.resultContent"
        )

        for card in cards:
            p = Prospect(job_portal="indeed.pl")

            title_el = (
                card.select_one("h2.jobTitle a")
                or card.select_one("a[data-jk]")
                or card.select_one("a[id^='job_']")
            )
            if title_el:
                p.job_title = title_el.get_text(strip=True)
                href = title_el.get("href", "")
                if href and not href.startswith("http"):
                    href = urljoin("https://pl.indeed.com", href)
                p.job_url = href

            company_el = (
                card.select_one("span[data-testid='company-name']")
                or card.select_one("span.companyName")
                or card.select_one("span.company")
            )
            if company_el:
                p.company_name = company_el.get_text(strip=True)

            loc_el = (
                card.select_one("div[data-testid='text-location']")
                or card.select_one("div.companyLocation")
            )
            if loc_el:
                p.city = loc_el.get_text(strip=True).split(",")[0].strip()

            date_el = card.select_one("span.date, span[class*='date']")
            if date_el:
                p.job_date = date_el.get_text(strip=True)

            if p.company_name and not is_excluded(p.company_name):
                prospects.append(p)

        log.info("  page %d → %d offers so far", page + 1, len(prospects))

    return prospects


# ---------------------------------------------------------------------------
# Scraper: Panorama Firm
# ---------------------------------------------------------------------------

def scrape_panoramafirm(position: str, max_pages: int = 2) -> list[Prospect]:
    """Scrape panoramafirm.pl for companies related to a position keyword."""
    prospects = []
    query = quote_plus(position)
    base = f"https://panoramafirm.pl/szukaj?k={query}"

    log.info("PanoramaFirm: searching '%s'", position)

    for page in range(1, max_pages + 1):
        page_url = f"{base}&p={page}" if page > 1 else base
        soup = fetch(page_url)
        if not soup:
            break

        cards = soup.select(
            "div.company-item, "
            "div.search-result, "
            "div[class*='company'], "
            "li.search-results__item"
        )

        for card in cards:
            p = Prospect(job_portal="panoramafirm.pl")

            name_el = card.select_one(
                "h2 a, h3 a, a.company-name, a[class*='name']"
            )
            if name_el:
                p.company_name = name_el.get_text(strip=True)
                href = name_el.get("href", "")
                if href and not href.startswith("http"):
                    href = urljoin("https://panoramafirm.pl", href)
                p.job_url = href

            # Try to get website, phone, email from the card
            page_text = card.get_text(" ", strip=True)
            emails = extract_emails(page_text)
            phones = extract_phones(page_text)
            if emails:
                p.email = emails[0]
            if phones:
                p.phone = phones[0]

            addr_el = card.select_one(
                "span.address, div.address, span[class*='address']"
            )
            if addr_el:
                p.city = addr_el.get_text(strip=True).split(",")[0].strip()

            if p.company_name and not is_excluded(p.company_name):
                prospects.append(p)

        log.info("  page %d → %d entries so far", page, len(prospects))

    return prospects


# ---------------------------------------------------------------------------
# Scraper: Rejestr.io
# ---------------------------------------------------------------------------

def scrape_rejestr(company_name: str) -> dict:
    """Look up company on rejestr.io to get NIP, KRS, people."""
    query = quote_plus(company_name)
    url = f"https://rejestr.io/szukaj?q={query}"
    soup = fetch(url)
    info: dict = {}
    if not soup:
        return info

    # First result link
    first = soup.select_one("a[href*='/krs/']")
    if not first:
        return info

    detail_url = first.get("href", "")
    if detail_url and not detail_url.startswith("http"):
        detail_url = urljoin("https://rejestr.io", detail_url)

    detail = fetch(detail_url)
    if not detail:
        return info

    page_text = detail.get_text(" ", strip=True)

    nip_match = NIP_RE.search(page_text)
    if nip_match:
        info["nip"] = re.sub(r"\D", "", nip_match.group(1))

    krs_match = KRS_RE.search(page_text)
    if krs_match:
        info["krs"] = krs_match.group(1)

    # Look for board members / management
    people = []
    mgmt_section = detail.find(string=re.compile(r"Zarząd|Reprezentacja|Organ"))
    if mgmt_section:
        parent = mgmt_section.find_parent("div") or mgmt_section.find_parent("section")
        if parent:
            rows = parent.find_all(["li", "tr", "div"])
            for row in rows:
                text = row.get_text(" ", strip=True)
                for prio, titles in DECISION_MAKER_TITLES_PRIORITY.items():
                    for t in titles:
                        if t.lower() in text.lower():
                            # Try to extract name
                            name_parts = re.findall(r"[A-ZĄĆĘŁŃÓŚŹŻ][a-ząćęłńóśźż]+", text)
                            if len(name_parts) >= 2:
                                people.append({
                                    "name": " ".join(name_parts[:2]),
                                    "title": t.title(),
                                    "priority": prio,
                                })

    if people:
        people.sort(key=lambda x: x["priority"])
        info["people"] = people

    emails = extract_emails(page_text)
    phones = extract_phones(page_text)
    if emails:
        info["emails"] = emails
    if phones:
        info["phones"] = phones

    return info


# ---------------------------------------------------------------------------
# Company website enrichment
# ---------------------------------------------------------------------------

def enrich_from_website(prospect: Prospect) -> None:
    """Try to extract contact info from the company website."""
    if not prospect.website:
        return

    base = prospect.website
    if not base.startswith("http"):
        base = "https://" + base

    # Try common contact pages
    contact_paths = ["", "/kontakt", "/contact", "/o-nas", "/about", "/zespol", "/team"]

    for path in contact_paths:
        url = base.rstrip("/") + path
        soup = fetch(url, retries=1)
        if not soup:
            continue

        page_text = soup.get_text(" ", strip=True)

        if not prospect.email:
            emails = extract_emails(page_text)
            if emails:
                prospect.email = emails[0]

        if not prospect.phone:
            phones = extract_phones(page_text)
            if phones:
                prospect.phone = phones[0]

        # Look for decision makers on the page
        if not prospect.decision_maker_name:
            for prio, titles in DECISION_MAKER_TITLES_PRIORITY.items():
                for title in titles:
                    pattern = re.compile(
                        rf"({title})[:\s\-–—]+([A-ZĄĆĘŁŃÓŚŹŻ][a-ząćęłńóśźż]+ [A-ZĄĆĘŁŃÓŚŹŻ][a-ząćęłńóśźż]+)",
                        re.IGNORECASE,
                    )
                    match = pattern.search(page_text)
                    if match:
                        prospect.decision_maker_title = match.group(1).strip()
                        prospect.decision_maker_name = match.group(2).strip()
                        prospect.decision_maker_priority = prio
                        break
                if prospect.decision_maker_name:
                    break

        if prospect.email and prospect.phone and prospect.decision_maker_name:
            break  # got everything we need


# ---------------------------------------------------------------------------
# Google search for company website
# ---------------------------------------------------------------------------

def find_company_website(company_name: str) -> str:
    """Try to find company website via a simple search on panoramafirm."""
    query = quote_plus(company_name + " Polska")
    url = f"https://panoramafirm.pl/szukaj?k={query}"
    soup = fetch(url, retries=1)
    if not soup:
        return ""

    # Look for external website link in the first result
    link = soup.select_one("a[href*='http'][rel='nofollow'], a.website-link")
    if link:
        href = link.get("href", "")
        parsed = urlparse(href)
        if parsed.scheme and parsed.netloc:
            if "panoramafirm" not in parsed.netloc:
                return href
    return ""


# ---------------------------------------------------------------------------
# Deduplication and merging
# ---------------------------------------------------------------------------

def normalize_company(name: str) -> str:
    """Normalize company name for dedup."""
    name = name.lower().strip()
    for suffix in [" sp. z o.o.", " s.a.", " sp.j.", " sp.k.", " spółka z o.o.",
                   " sp. z o. o.", " spółka akcyjna", " s.c."]:
        name = name.replace(suffix, "")
    name = re.sub(r"\s+", " ", name).strip()
    return name


def deduplicate(prospects: list[Prospect]) -> list[Prospect]:
    """Remove duplicate companies, keeping the most complete entry."""
    seen: dict[str, Prospect] = {}
    for p in prospects:
        key = normalize_company(p.company_name)
        if not key:
            continue
        if key in seen:
            existing = seen[key]
            # Merge: keep more complete data
            if not existing.email and p.email:
                existing.email = p.email
            if not existing.phone and p.phone:
                existing.phone = p.phone
            if not existing.website and p.website:
                existing.website = p.website
            if not existing.decision_maker_name and p.decision_maker_name:
                existing.decision_maker_name = p.decision_maker_name
                existing.decision_maker_title = p.decision_maker_title
                existing.decision_maker_priority = p.decision_maker_priority
            if not existing.city and p.city:
                existing.city = p.city
            if not existing.industry and p.industry:
                existing.industry = p.industry
        else:
            seen[key] = p
    return list(seen.values())


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="06423A", end_color="06423A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ACTIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COLUMNS = [
    ("Lp.", 5),
    ("Nazwa firmy", 30),
    ("Branża", 20),
    ("Strona www", 30),
    ("Miasto", 18),
    ("Osoba decyzyjna", 25),
    ("Stanowisko osoby", 25),
    ("Email", 30),
    ("Telefon", 18),
    ("Stanowisko z ogłoszenia", 30),
    ("Portal", 18),
    ("Data publikacji", 16),
    ("Link do ogłoszenia", 40),
    ("NIP", 14),
    ("KRS", 14),
    ("Aktywne ogłoszenie", 12),
]


def write_excel(prospects: list[Prospect], output_path: str) -> None:
    """Write prospects to a formatted Excel workbook."""
    wb = Workbook()

    # ---- Main sheet ----
    ws = wb.active
    ws.title = "Prospekty B2B"

    # Headers
    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, p in enumerate(prospects, 2):
        values = [
            row_idx - 1,
            p.company_name,
            p.industry,
            p.website,
            p.city,
            p.decision_maker_name,
            p.decision_maker_title,
            p.email,
            p.phone,
            p.job_title,
            p.job_portal,
            p.job_date,
            p.job_url,
            p.nip,
            p.krs,
            "TAK" if p.is_active else "NIE",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if p.is_active:
                cell.fill = ACTIVE_FILL

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(prospects) + 1}"

    # Freeze top row
    ws.freeze_panes = "A2"

    # ---- Statistics sheet ----
    ws_stats = wb.create_sheet("Statystyki")

    stats_headers = [
        ("Statystyka", 35),
        ("Wartość", 20),
    ]
    for col_idx, (header, width) in enumerate(stats_headers, 1):
        cell = ws_stats.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws_stats.column_dimensions[get_column_letter(col_idx)].width = width

    # Compute stats
    total = len(prospects)
    with_email = sum(1 for p in prospects if p.email)
    with_phone = sum(1 for p in prospects if p.phone)
    with_dm = sum(1 for p in prospects if p.decision_maker_name)
    with_website = sum(1 for p in prospects if p.website)
    active_offers = sum(1 for p in prospects if p.is_active)

    portal_counts = Counter(p.job_portal for p in prospects if p.job_portal)
    industry_counts = Counter(p.industry for p in prospects if p.industry)
    city_counts = Counter(p.city for p in prospects if p.city)
    position_counts = Counter(p.job_title for p in prospects if p.job_title)

    stats_data = [
        ("Łączna liczba prospektów", total),
        ("Z adresem email", f"{with_email} ({with_email*100//max(total,1)}%)"),
        ("Z numerem telefonu", f"{with_phone} ({with_phone*100//max(total,1)}%)"),
        ("Z osobą decyzyjną", f"{with_dm} ({with_dm*100//max(total,1)}%)"),
        ("Ze stroną www", f"{with_website} ({with_website*100//max(total,1)}%)"),
        ("Aktywne ogłoszenia", active_offers),
        ("Data wygenerowania", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("--- Podział wg portalu ---", ""),
    ]
    for portal, count in portal_counts.most_common():
        stats_data.append((f"  {portal}", count))

    stats_data.append(("", ""))
    stats_data.append(("--- Podział wg branży ---", ""))
    for ind, count in industry_counts.most_common(10):
        stats_data.append((f"  {ind}", count))

    stats_data.append(("", ""))
    stats_data.append(("--- Top 10 miast ---", ""))
    for city, count in city_counts.most_common(10):
        stats_data.append((f"  {city}", count))

    stats_data.append(("", ""))
    stats_data.append(("--- Top 10 stanowisk z ogłoszeń ---", ""))
    for pos, count in position_counts.most_common(10):
        stats_data.append((f"  {pos[:60]}", count))

    for row_idx, (label, value) in enumerate(stats_data, 2):
        cell_a = ws_stats.cell(row=row_idx, column=1, value=label)
        cell_b = ws_stats.cell(row=row_idx, column=2, value=value)
        cell_a.border = THIN_BORDER
        cell_b.border = THIN_BORDER
        if str(label).startswith("---"):
            cell_a.font = Font(bold=True)

    wb.save(output_path)
    log.info("Excel saved: %s (%d prospects)", output_path, total)


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def run_pipeline(output_path: str, min_prospects: int = 50) -> None:
    """Execute the full scraping pipeline."""
    all_prospects: list[Prospect] = []

    # Phase 1: Scrape job portals
    log.info("=" * 60)
    log.info("PHASE 1: Scraping job portals")
    log.info("=" * 60)

    for position in SEARCH_POSITIONS:
        log.info("--- Position: %s ---", position)

        # Pracuj.pl
        try:
            results = scrape_pracuj(position, max_pages=2)
            all_prospects.extend(results)
            log.info("  Pracuj.pl: %d results", len(results))
        except Exception as exc:
            log.error("  Pracuj.pl error: %s", exc)

        # Indeed.pl
        try:
            results = scrape_indeed(position, max_pages=2)
            all_prospects.extend(results)
            log.info("  Indeed.pl: %d results", len(results))
        except Exception as exc:
            log.error("  Indeed.pl error: %s", exc)

        # PanoramaFirm
        try:
            results = scrape_panoramafirm(position, max_pages=1)
            all_prospects.extend(results)
            log.info("  PanoramaFirm: %d results", len(results))
        except Exception as exc:
            log.error("  PanoramaFirm error: %s", exc)

    log.info("Total raw results: %d", len(all_prospects))

    # Phase 2: Deduplicate
    log.info("=" * 60)
    log.info("PHASE 2: Deduplication")
    log.info("=" * 60)
    all_prospects = deduplicate(all_prospects)
    log.info("After dedup: %d unique companies", len(all_prospects))

    # Phase 3: Enrich with company data
    log.info("=" * 60)
    log.info("PHASE 3: Enrichment (websites, contacts, KRS)")
    log.info("=" * 60)

    for i, p in enumerate(all_prospects):
        if i >= min_prospects * 2:
            # Don't enrich too many – focus on first batch
            break

        log.info("  [%d/%d] Enriching: %s", i + 1, len(all_prospects), p.company_name)

        # Find website if missing
        if not p.website:
            try:
                p.website = find_company_website(p.company_name)
            except Exception as exc:
                log.warning("    Website lookup failed: %s", exc)

        # Enrich from website
        try:
            enrich_from_website(p)
        except Exception as exc:
            log.warning("    Website enrichment failed: %s", exc)

        # Look up in rejestr.io
        try:
            reg_info = scrape_rejestr(p.company_name)
            if reg_info.get("nip"):
                p.nip = reg_info["nip"]
            if reg_info.get("krs"):
                p.krs = reg_info["krs"]
            if not p.email and reg_info.get("emails"):
                p.email = reg_info["emails"][0]
            if not p.phone and reg_info.get("phones"):
                p.phone = reg_info["phones"][0]
            if not p.decision_maker_name and reg_info.get("people"):
                best = reg_info["people"][0]
                p.decision_maker_name = best["name"]
                p.decision_maker_title = best["title"]
                p.decision_maker_priority = best["priority"]
        except Exception as exc:
            log.warning("    Rejestr.io lookup failed: %s", exc)

    # Phase 4: Final filtering
    log.info("=" * 60)
    log.info("PHASE 4: Final filtering")
    log.info("=" * 60)

    # Remove excluded industries once more (in case enrichment revealed the industry)
    filtered = [p for p in all_prospects if not is_excluded(
        f"{p.company_name} {p.industry} {p.job_title}"
    )]
    log.info("After final filter: %d prospects", len(filtered))

    if len(filtered) < min_prospects:
        log.warning(
            "Only found %d prospects (target: %d). "
            "Some portals may have blocked scraping or changed their HTML structure. "
            "Consider running again or checking the portals manually.",
            len(filtered),
            min_prospects,
        )

    # Phase 5: Write output
    log.info("=" * 60)
    log.info("PHASE 5: Writing Excel")
    log.info("=" * 60)

    # Set today's date for those without a date
    today_str = date.today().strftime("%Y-%m-%d")
    for p in filtered:
        if not p.job_date:
            p.job_date = today_str

    write_excel(filtered, output_path)

    # Summary
    log.info("=" * 60)
    log.info("DONE! Summary:")
    log.info("  Total prospects: %d", len(filtered))
    log.info("  With email: %d", sum(1 for p in filtered if p.email))
    log.info("  With phone: %d", sum(1 for p in filtered if p.phone))
    log.info("  With decision maker: %d", sum(1 for p in filtered if p.decision_maker_name))
    log.info("  Output file: %s", output_path)
    log.info("=" * 60)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="B2B Prospect Database Builder – scrapes Polish job portals",
    )
    parser.add_argument(
        "--output", "-o",
        default=f"prospekty_b2b_{date.today().strftime('%Y%m%d')}.xlsx",
        help="Output Excel file path (default: prospekty_b2b_YYYYMMDD.xlsx)",
    )
    parser.add_argument(
        "--min-prospects", "-n",
        type=int,
        default=50,
        help="Minimum number of prospects to collect (default: 50)",
    )
    args = parser.parse_args()

    log.info("Starting B2B Prospect Scraper")
    log.info("Output: %s | Target: %d+ prospects", args.output, args.min_prospects)

    run_pipeline(args.output, args.min_prospects)


if __name__ == "__main__":
    main()
