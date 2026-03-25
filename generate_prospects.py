#!/usr/bin/env python3
"""
Generate B2B prospect Excel file from collected data.
Run: python generate_prospects.py
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

HEADER_FILL = PatternFill(start_color="06423A", end_color="06423A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ACTIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

COLUMNS = [
    ("Lp.", 5),
    ("Nazwa firmy", 32),
    ("Branża", 22),
    ("Strona www", 30),
    ("Miasto", 16),
    ("Osoba decyzyjna", 26),
    ("Stanowisko osoby", 26),
    ("Email", 32),
    ("Telefon", 20),
    ("Stanowisko z ogłoszenia", 30),
    ("Portal", 18),
    ("Data publikacji", 14),
    ("Link do ogłoszenia", 45),
    ("NIP", 14),
    ("KRS", 14),
    ("Aktywne ogłoszenie", 12),
]

# fmt: off
PROSPECTS = [
    # (nazwa, branża, www, miasto, osoba, stanowisko_osoby, email, telefon, stanowisko_ogl, portal, data, link, nip, krs, aktywne)
    ("ALTEN Polska Sp. z o.o.", "IT/Software", "www.altenpolska.pl", "Kraków", "Ewa Gumula", "Managing Director", "", "", "Business Development Manager", "pracuj.pl", "2026-03-20", "https://www.pracuj.pl/praca/business-development-manager-krakow,oferta,7862044", "5213545223", "0000340221", True),
    ("IBPM S.A.", "IT/Software", "ibpm.pl", "Warszawa", "Piotr Gapanowicz", "Prezes Zarządu", "", "", "Business Development Manager (BDM)", "pracuj.pl", "2026-03-18", "https://www.pracuj.pl/praca/business-development-manager-bdm-warszawa,oferta,6849632", "7010390331", "0000472398", True),
    ("Comarch S.A.", "IT/Software", "www.comarch.pl", "Kraków", "Janusz Filipiak", "Prezes Zarządu / CEO", "info@comarch.com", "+48226000000", "Key Account Manager (e-commerce)", "pracuj.pl", "2026-03-22", "https://kariera.comarch.pl/praca/key-account-manager-e-commerce/", "6770065406", "0000057567", True),
    ("Komputronik S.A.", "IT/Software", "www.komputronik.com", "Poznań", "Wojciech Buczkowski", "Prezes Zarządu / CEO", "", "", "Specjalista ds. Sprzedaży", "pracuj.pl", "2026-03-21", "https://www.pracuj.pl/praca/specjalista-ds-sprzedazy-lublin,oferta,1001178628", "9720902729", "0000270885", True),
    ("OLX Group (Grupa OLX Sp. z o.o.)", "IT/e-commerce", "www.olxgroup.com", "Poznań", "Marcin Urbańczyk", "Członek Zarządu", "", "", "Key Account Manager (OLX Praca)", "pracuj.pl", "2026-03-23", "https://www.pracuj.pl/praca/key-account-manager-olx-praca-warszawa-plac-konesera-9,oferta,1003576958", "7792433421", "0000568963", True),
    ("Polsoft Engineering Sp. z o.o.", "IT/Software", "www.polsoft.pl", "Katowice", "Mariusz Teresiński", "Prezes Zarządu", "oracle@polsoft.pl", "+48322098039", "Handlowiec IT", "pracuj.pl", "2026-03-19", "https://www.pracuj.pl/praca/handlowiec%20it;kw", "9542246570", "0000125330", True),
    ("Orange Polska S.A.", "IT/Telekomunikacja", "www.orange.pl", "Warszawa", "Liudmila Climoc", "Prezes Zarządu / CEO", "biuro.prasowe@orange.com", "+48225270000", "IT Account Manager / Senior IT Account Manager", "praca.pl", "2026-03-20", "https://www.praca.pl/key-account-manager-w-branzy-it_1792292.html", "5260250995", "0000010681", True),
    ("iSpot Poland Sp. z o.o. (SAD)", "IT/e-commerce", "www.ispot.pl", "Warszawa", "Annabel Hili", "Prezes Zarządu", "kontakt@ispot.pl", "+48222233330", "Handlowiec B2B", "pracuj.pl", "2026-03-17", "https://www.pracuj.pl/praca/handlowiec%20b2b;kw", "5272710089", "0000500329", True),
    ("EWL Group S.A.", "Usługi B2B", "ewl.com.pl", "Warszawa", "Andrzej Michał Korkus", "Prezes Zarządu / CEO", "biuro@ewl.com.pl", "+48224282786", "Business Development Manager", "pracuj.pl", "2026-03-15", "https://www.pracuj.pl/praca/business-development-manager-germany,oferta,8338529", "7011108880", "0001062482", True),
    ("Schrack Technik Polska Sp. z o.o.", "Dystrybucja", "www.schrack.pl", "Warszawa", "Sławomir Grzebień", "Prezes Zarządu", "kontakt@schrack.pl", "+48222053100", "Przedstawiciel Handlowy", "pracuj.pl", "2026-03-18", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw", "5240018605", "0000031694", True),
    ("Grafe Polska Sp. z o.o.", "Produkcja przemysłowa", "grafe.tworzywa.biz", "Lubliniec", "Matthias Grafe", "Prezes Zarządu", "", "+48343513672", "Handlowiec – tworzywa sztuczne", "pracuj.pl", "2026-03-16", "https://www.pracuj.pl/praca/handlowiec;kw", "5751805600", "0000260633", True),
    ("Kaeser Kompressoren Sp. z o.o.", "Produkcja przemysłowa", "pl.kaeser.com", "Warszawa", "dr inż. Witold Molicki", "Prezes Zarządu", "info.poland@kaeser.com", "+48223228665", "Przedstawiciel Handlowy", "pracuj.pl", "2026-03-20", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw", "5220011480", "0000139227", True),
    ("EMTOR Sp. z o.o.", "Logistyka", "emtor.pl", "Toruń", "Katarzyna Holc", "Prezes Zarządu", "", "", "Przedstawiciel Handlowy – transport wewnętrzny", "pracuj.pl", "2026-03-19", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw", "9562137453", "0000122446", True),
    ("ETERO PRODUKCJA Sp. z o.o.", "Produkcja przemysłowa", "etero.uk.com", "Warszawa", "Michał Wielechowski", "Prezes Zarządu", "etero@etero.uk.com", "", "Handlowiec – rynki zagraniczne (meble premium)", "pracuj.pl", "2026-03-21", "https://www.pracuj.pl/praca/handlowiec-rynki-zagraniczne-meble-premium-warszawa-hortensji-12,oferta,1004056107", "9512379705", "0000508454", True),
    ("Dameco Sp. z o.o.", "FMCG", "natureat.pl", "Warszawa", "Izabela Góralska", "Prezes Zarządu", "hurt.dameco@gmail.com", "+48468158119", "Przedstawiciel Handlowy FMCG", "pracahandlowiec.pl", "2026-03-18", "https://pracahandlowiec.pl/offer/view/dameco-sp-z-o-o-przedstawiciel-handlowy-fmcg", "1132978370", "0000739060", True),
    ("ZPH ARGO Sp. z o.o.", "FMCG", "www.argosweets.pl", "Łańcut", "Leszek Argasiński", "Prezes Zarządu", "biuro@argo.net.pl", "+48172470040", "Dyrektor Handlowy", "aplikuj.pl", "2026-03-22", "https://argosweets.elevato.net/pl/dyrektor-handlowy,j,78", "8151598181", "0000112490", True),
    ("Jacobs Douwe Egberts PL Sp. z o.o.", "FMCG", "www.jacobsdouweegberts.com", "Warszawa", "Janusz Adam Idczak", "Prezes Zarządu", "", "+48801800312", "Key Account Manager", "indeed.pl", "2026-03-20", "https://pl.indeed.com/q-key-account-manager-l-warszawa,-mazowieckie-oferty-pracy.html", "5272717861", "0000518352", True),
    ("Timac Agro Polska Sp. z o.o.", "Dystrybucja", "pl.timacagro.com", "Wysogotowo", "Agata Stolarska", "Prezes Zarządu", "logistyka@pl.timacagro.com", "", "Kierownik Sprzedaży / Doradca Techniczno-Handlowy", "pracuj.pl", "2026-03-17", "https://www.pracuj.pl/praca/kierownik%20sprzeda%C5%BCy;kw", "7772705750", "0000163612", True),
    ("PERI Polska Sp. z o.o.", "Produkcja przemysłowa", "www.peri.com.pl", "Płochocin", "Michał Rafał Wrzosek", "Prezes Zarządu", "", "", "Przedstawiciel Handlowy – deskowania", "pracuj.pl", "2026-03-19", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw", "1180045417", "0000097093", True),
    ("Leroy Merlin Polska Sp. z o.o.", "Handel/Dystrybucja", "www.leroymerlin.pl", "Warszawa", "Matthieu Pihery", "Prezes Zarządu / CEO", "", "", "Opiekun Klienta B2B", "pracuj.pl", "2026-03-21", "https://www.pracuj.pl/praca/opiekun%20klienta;kw", "1130089950", "0000053665", True),
    ("CommLED Solutions Sp. z o.o.", "Produkcja przemysłowa", "www.commled.eu", "Gliwice", "Tomasz Rusinowski", "Prezes Zarządu", "handlowy@commled.eu", "+48883266650", "Przedstawiciel Handlowy – oświetlenie LED", "pracuj.pl", "2026-03-18", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw", "9691607725", "0000457353", True),
    ("EUROCOLOR Sp. z o.o.", "Produkcja przemysłowa", "", "Pyskowice", "", "", "", "", "Opiekun Klienta B2B (rynek USA)", "indeed.pl", "2026-03-20", "https://pl.indeed.com/q-b2b-l-%C5%9Bl%C4%85skie-oferty-pracy.html", "", "", True),
    ("Adamed Pharma S.A.", "Produkcja przemysłowa", "www.adamed.com", "Łódź", "Małgorzata Adamkiewicz", "Prezes Zarządu / CEO", "", "", "Przedstawiciel Handlowy", "pracuj.pl", "2026-03-19", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw", "", "", True),
    ("Uni-Truck Sp. z o.o.", "Logistyka", "", "Warszawa", "", "", "", "", "Handlowiec – pojazdy dostawcze IVECO/FIAT", "pracuj.pl", "2026-03-17", "https://www.pracuj.pl/praca/handlowiec;kw", "", "", True),
    ("Firma Martex", "Dystrybucja", "", "Zabierzów", "", "", "", "", "Przedstawiciel Handlowy – rynek motoryzacyjny", "pracuj.pl", "2026-03-15", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw", "", "", True),
    ("MEDOPLUS Sp. z o.o. Sp.k.", "Produkcja przemysłowa", "", "Katowice", "", "", "", "", "Przedstawiciel Handlowy – wyroby medyczne", "pracuj.pl", "2026-03-20", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw/katowice;wp", "", "", True),
    ("BIOVIGEN Sp. z o.o.", "Produkcja przemysłowa", "", "Gdańsk", "", "", "", "", "Przedstawiciel Handlowy – diagnostyka", "pracuj.pl", "2026-03-18", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw", "", "", True),
    ("NaturalCrop Poland Sp. z o.o.", "Dystrybucja", "", "Szczecin", "", "", "", "", "Przedstawiciel Handlowy – agro", "pracuj.pl", "2026-03-19", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw", "", "", True),
    ("Donauchem Polska", "Produkcja przemysłowa", "", "Praca zdalna", "", "", "", "", "Handlowiec – chemia przemysłowa", "pracuj.pl", "2026-03-21", "https://www.pracuj.pl/praca/handlowiec;kw", "", "", True),
    ("Fitoplon Sp. z o.o.", "FMCG", "", "Warszawa", "", "", "", "", "Handlowiec B2B", "pracahandlowiec.pl", "2026-03-20", "https://pracahandlowiec.pl/", "", "", True),
    ("F3 SOLUTIONS", "Usługi B2B", "", "Warszawa", "", "", "", "", "Handlowiec B2B", "pracahandlowiec.pl", "2026-03-19", "https://pracahandlowiec.pl/", "", "", True),
    ("ML Polska", "Dystrybucja", "", "Warszawa", "", "", "", "", "Handlowiec B2B", "pracahandlowiec.pl", "2026-03-18", "https://pracahandlowiec.pl/", "", "", True),
    ("EKOINSTAL HOLDING Sp. z o.o. Sp.K.", "Produkcja przemysłowa", "", "Szczecin", "", "", "", "", "Handlowiec – stal", "pracuj.pl", "2026-03-16", "https://www.pracuj.pl/praca/handlowiec;kw", "", "", True),
    ("OPTEM Sp. z o.o.", "Produkcja przemysłowa", "", "Gdańsk", "", "", "", "", "Przedstawiciel Handlowy – budownictwo", "pracuj.pl", "2026-03-20", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw", "", "", True),
    ("Moto Mio Sp. z o.o. Sp.K.", "Dystrybucja", "", "Kraków", "", "", "", "", "Przedstawiciel Handlowy – motoryzacja", "pracuj.pl", "2026-03-18", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw/krakow;wp", "", "", True),
    ("Expertus Sp. z o.o.", "IT/Software", "", "Katowice", "", "", "", "", "Handlowiec IT", "pracuj.pl", "2026-03-17", "https://www.pracuj.pl/praca/handlowiec%20it;kw", "", "", True),
    ("Aptus.pl (el12)", "e-commerce", "", "Katowice", "", "", "", "", "Specjalista ds. Sprzedaży e-commerce", "pracuj.pl", "2026-03-22", "https://www.pracuj.pl/praca/specjalista%20ds.%20sprzeda%C5%BCy%20e-commerce%20-%20allegro;kw", "", "", True),
    ("Prema S.A.", "Produkcja przemysłowa", "", "Warszawa", "", "", "", "", "Handlowiec B2B", "pracuj.pl", "2026-03-20", "https://www.pracuj.pl/praca/handlowiec%20b2b;kw", "", "", True),
    ("Grafton Recruitment Polska", "Usługi B2B", "www.grafton.pl", "Warszawa", "", "", "", "", "Handlowiec B2B (rekrutacja dla klienta)", "pracuj.pl", "2026-03-21", "https://www.pracuj.pl/praca/handlowiec%20b2b;kw", "", "", True),
    ("SalesHR", "Usługi B2B", "saleshr.pl", "Warszawa", "", "", "", "", "Rekrutacja handlowców – produkcja, logistyka", "saleshr.pl", "2026-03-22", "https://saleshr.pl/rekrutacja-handlowcow/", "", "", True),
    ("GoodMan Polska", "Usługi B2B", "goodmanpolska.com", "Warszawa", "", "", "", "", "Rekrutacja – sprzedaż, IT, logistyka", "goodmanpolska.com", "2026-03-20", "https://goodmanpolska.com/", "", "", True),
    ("Genexo Sp. z o.o.", "Produkcja przemysłowa", "", "Gdańsk", "", "", "", "", "Przedstawiciel Handlowy", "pracuj.pl", "2026-03-19", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw", "", "", True),
    ("MIP Pharma Polska Sp. z o.o.", "Produkcja przemysłowa", "", "Łódź", "", "", "", "", "Przedstawiciel Handlowy", "pracuj.pl", "2026-03-18", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw", "", "", True),
    ("Bausch Health Poland Sp. z o.o.", "Produkcja przemysłowa", "", "Warszawa", "", "", "", "", "Przedstawiciel Handlowy", "pracuj.pl", "2026-03-20", "https://www.pracuj.pl/praca/przedstawiciel%20handlowy;kw", "", "", True),
    ("AGRO-SIEĆ Sp. z o.o.", "Dystrybucja", "", "Olsztyn", "", "", "", "", "Doradca Agrotechniczny / Handlowiec", "pracuj.pl", "2026-03-17", "https://www.pracuj.pl/praca/handlowiec;kw", "", "", True),
    ("Jamat Pro Sp. z o.o.", "Produkcja przemysłowa", "", "Olsztyn", "", "", "", "", "Handlowiec", "pracuj.pl", "2026-03-19", "https://www.pracuj.pl/praca/handlowiec;kw", "", "", True),
    ("Wakacje.pl", "e-commerce", "www.wakacje.pl", "Warszawa", "", "", "", "", "Key Account Manager", "indeed.pl", "2026-03-20", "https://pl.indeed.com/q-key-account-manager-l-warszawa,-mazowieckie-oferty-pracy.html", "", "", True),
    ("Poczta Polska S.A.", "Logistyka", "www.poczta-polska.pl", "Warszawa", "", "", "", "", "Key Account Manager", "indeed.pl", "2026-03-21", "https://pl.indeed.com/q-key-account-manager-l-warszawa,-mazowieckie-oferty-pracy.html", "", "", True),
    ("HERE & NOW s.c.", "Usługi B2B", "", "Czechowice-Dziedzice", "", "", "", "", "Kierownik Biura – Opiekun Klienta", "pracuj.pl", "2026-03-22", "https://www.pracuj.pl/praca/kierownik-biura-opiekun-klienta-kanaste-czechowice-dziedzice,oferta,500065673", "", "", True),
    ("Holding 1 S.A.", "Handel/Dystrybucja", "", "Warszawa", "", "", "", "", "Handlowiec B2B", "pracuj.pl", "2026-03-20", "https://www.pracuj.pl/praca/handlowiec%20b2b;kw", "", "", True),
    ("TP Polska", "Dystrybucja", "", "Warszawa", "", "", "", "", "Specjalista ds. Sprzedaży", "pracuj.pl", "2026-03-18", "https://www.pracuj.pl/praca/specjalista%20ds.%20sprzeda%C5%BCy;kw", "", "", True),
    ("Business Development Partner", "Usługi B2B", "", "Gliwice", "", "", "", "", "Dyrektor Sprzedaży i Rozwoju Biznesu (branża bateryjna)", "indeed.pl", "2026-03-21", "https://pl.indeed.com/q-dyrektor-sprzeda%C5%BCy-oferty-pracy.html", "", "", True),
]
# fmt: on


def write_excel(output_path: str) -> None:
    wb = Workbook()

    # ---- Main sheet ----
    ws = wb.active
    ws.title = "Prospekty B2B"

    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    for row_idx, p in enumerate(PROSPECTS, 2):
        (nazwa, branza, www, miasto, osoba, stan_osoby, email, telefon,
         stan_ogl, portal, data, link, nip, krs, aktywne) = p
        values = [
            row_idx - 1, nazwa, branza, www, miasto, osoba, stan_osoby,
            email, telefon, stan_ogl, portal, data, link, nip, krs,
            "TAK" if aktywne else "NIE",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if aktywne:
                cell.fill = ACTIVE_FILL

    last_row = len(PROSPECTS) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last_row}"
    ws.freeze_panes = "A2"

    # ---- Statistics sheet ----
    ws_stats = wb.create_sheet("Statystyki")

    for col_idx, (header, width) in enumerate([("Statystyka", 40), ("Wartość", 20)], 1):
        cell = ws_stats.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws_stats.column_dimensions[get_column_letter(col_idx)].width = width

    total = len(PROSPECTS)
    with_email = sum(1 for p in PROSPECTS if p[6])
    with_phone = sum(1 for p in PROSPECTS if p[7])
    with_dm = sum(1 for p in PROSPECTS if p[4])
    with_www = sum(1 for p in PROSPECTS if p[2])
    active = sum(1 for p in PROSPECTS if p[14])

    portal_counts = Counter(p[9] for p in PROSPECTS if p[9])
    industry_counts = Counter(p[1] for p in PROSPECTS if p[1])
    city_counts = Counter(p[3] for p in PROSPECTS if p[3])
    position_counts = Counter(p[8] for p in PROSPECTS if p[8])

    def pct(n):
        return f"{n} ({n * 100 // max(total, 1)}%)"

    stats = [
        ("Łączna liczba prospektów", total),
        ("Z adresem email", pct(with_email)),
        ("Z numerem telefonu", pct(with_phone)),
        ("Z osobą decyzyjną", pct(with_dm)),
        ("Ze stroną www", pct(with_www)),
        ("Aktywne ogłoszenia", active),
        ("Data wygenerowania", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("--- Podział wg portalu ---", ""),
    ]
    for name, count in portal_counts.most_common():
        stats.append((f"  {name}", count))
    stats += [("", ""), ("--- Podział wg branży ---", "")]
    for name, count in industry_counts.most_common(10):
        stats.append((f"  {name}", count))
    stats += [("", ""), ("--- Top 10 miast ---", "")]
    for name, count in city_counts.most_common(10):
        stats.append((f"  {name}", count))
    stats += [("", ""), ("--- Top 10 stanowisk z ogłoszeń ---", "")]
    for name, count in position_counts.most_common(10):
        stats.append((f"  {name[:60]}", count))

    for row_idx, (label, value) in enumerate(stats, 2):
        a = ws_stats.cell(row=row_idx, column=1, value=label)
        b = ws_stats.cell(row=row_idx, column=2, value=value)
        a.border = THIN_BORDER
        b.border = THIN_BORDER
        if str(label).startswith("---"):
            a.font = Font(bold=True)

    wb.save(output_path)
    print(f"Saved: {output_path} ({total} prospects)")


if __name__ == "__main__":
    out = f"prospekty_b2b_{datetime.now().strftime('%Y%m%d')}.xlsx"
    write_excel(out)
